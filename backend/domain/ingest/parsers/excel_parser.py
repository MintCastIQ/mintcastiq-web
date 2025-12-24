import openpyxl
from pathlib import Path


class ExcelChecklistParser:
    """
    Parses a contributor-provided Excel (.xlsx) checklist file containing
    multiple sheets. Yields raw dict rows with sheet context included.
    """

    def __init__(self, file_path):
        self.file_path = Path(file_path)

    def parse(self):
        if not self.file_path.exists():
            raise FileNotFoundError(f"Checklist file not found: {self.file_path}")

        workbook = openpyxl.load_workbook(self.file_path, data_only=True)

        # Always take the last sheet
        sheet = workbook.worksheets[-1]

        # Skip sheets with no data rows
        if sheet.max_row < 2:
            return []

        # Extract headers from row 1
        headers = [cell.value for cell in sheet[1]]

        # Skip sheets with no headers
        if not any(headers):
            return []

        raw_rows = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            raw_rows.append(dict(zip(headers, row)))

        return raw_rows
            
